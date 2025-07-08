# from flask import Flask, render_template, request
# import joblib
# import numpy as np
# import pandas as pd

# app = Flask(__name__)

# # Load models
# breast_model = joblib.load("top8_breast_cancer_model.joblib")
# cervical_model = joblib.load("cervical_cancer_top8_model.joblib")
# colorectal_model = joblib.load("crc_risk_model.joblib")
# crc_sample = pd.read_csv("crc_dataset.csv").iloc[[0]].copy()

# @app.route("/")
# def index():
#     return render_template("index.html")

# @app.route("/breast-cancer", methods=["GET", "POST"])
# def breast():
#     prob = None
#     if request.method == "POST":
#         features = [
#             float(request.form["concavity_worst"]),
#             float(request.form["perimeter_worst"]),
#             float(request.form["concave_points_worst"]),
#             float(request.form["radius_worst"]),
#             float(request.form["radius_mean"]),
#             float(request.form["area_worst"]),
#             float(request.form["area_mean"]),
#             float(request.form["perimeter_mean"])
#         ]
#         prob = breast_model.predict_proba([features])[0][1]
#     return render_template("breast-cancer.html", prob=prob)

# @app.route("/cervical", methods=["GET", "POST"])
# def cervical():
#     prob = None
#     if request.method == "POST":
#         features = [
#             float(request.form["Age"]),
#             float(request.form["Number_of_sexual_partners"]),
#             float(request.form["First_sexual_intercourse"]),
#             int(request.form["STDs_HPV"]),
#             int(request.form["Dx_HPV"]),
#             int(request.form["STDs"]),
#             float(request.form["STDs_Number_of_diagnosis"]),
#             float(request.form["STDs_Time_since_last_diagnosis"])
#         ]
#         prob = cervical_model.predict_proba([features])[0][1]
#     return render_template("cervical.html", prob=prob)

# @app.route("/colorectal", methods=["GET", "POST"])
# def colorectal():
#     prob = None
#     if request.method == "POST":
#         age = float(request.form["Age"])
#         bmi = float(request.form["BMI"])
#         data = crc_sample.copy()
#         data["Age"] = age
#         data["BMI"] = bmi
#         prob = colorectal_model.predict_proba(data)[0][1]
#     return render_template("colorectal.html", prob=prob)

# if __name__ == "__main__":
#     app.run(debug=True)

from flask import Flask, render_template, request, redirect, url_for, session, flash
import joblib
import numpy as np
import pandas as pd
from datetime import datetime, date
import os
from werkzeug.security import generate_password_hash, check_password_hash
import json
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
app.secret_key = 'your-secret-key-here-change-this-to-secure-key'  # Change this to a secure secret key

# Load models
breast_model = joblib.load("top8_breast_cancer_model.joblib")
cervical_model = joblib.load("cervical_cancer_top8_model.joblib")
colorectal_model = joblib.load("crc_risk_model.joblib")
crc_sample = pd.read_csv("crc_dataset.csv").iloc[[0]].copy()

# Excel file paths
USERS_FILE = "users.xlsx"
PREDICTIONS_FILE = "predictions.xlsx"

def init_excel_files():
    """Initialize Excel files if they don't exist"""
    if not os.path.exists(USERS_FILE):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Users"
        worksheet.append(['username', 'email', 'password_hash', 'created_at'])
        workbook.save(USERS_FILE)
    
    if not os.path.exists(PREDICTIONS_FILE):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Predictions"
        worksheet.append(['username', 'prediction_type', 'date', 'time', 'input_data', 'probability', 'result'])
        workbook.save(PREDICTIONS_FILE)

def load_users():
    """Load users from Excel file"""
    try:
        return pd.read_excel(USERS_FILE)
    except:
        return pd.DataFrame(columns=['username', 'email', 'password_hash', 'created_at'])

def save_user(username, email, password):
    """Save new user to Excel file"""
    try:
        workbook = load_workbook(USERS_FILE)
        worksheet = workbook.active
        worksheet.append([username, email, generate_password_hash(password), datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        workbook.save(USERS_FILE)
    except:
        # If file doesn't exist, create new one
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(['username', 'email', 'password_hash', 'created_at'])
        worksheet.append([username, email, generate_password_hash(password), datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        workbook.save(USERS_FILE)

def load_predictions():
    """Load predictions from Excel file"""
    try:
        return pd.read_excel(PREDICTIONS_FILE)
    except:
        return pd.DataFrame(columns=['username', 'prediction_type', 'date', 'time', 'input_data', 'probability', 'result'])

def save_prediction(username, prediction_type, input_data, probability):
    """Save prediction to Excel file"""
    try:
        workbook = load_workbook(PREDICTIONS_FILE)
        worksheet = workbook.active
        now = datetime.now()
        result = "High Risk" if probability > 0.5 else "Low Risk"
        worksheet.append([
            username,
            prediction_type,
            now.strftime('%Y-%m-%d'),
            now.strftime('%H:%M:%S'),
            json.dumps(input_data),
            probability,
            result
        ])
        workbook.save(PREDICTIONS_FILE)
    except:
        # If file doesn't exist, create new one
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(['username', 'prediction_type', 'date', 'time', 'input_data', 'probability', 'result'])
        now = datetime.now()
        result = "High Risk" if probability > 0.5 else "Low Risk"
        worksheet.append([
            username,
            prediction_type,
            now.strftime('%Y-%m-%d'),
            now.strftime('%H:%M:%S'),
            json.dumps(input_data),
            probability,
            result
        ])
        workbook.save(PREDICTIONS_FILE)

def get_user_predictions(username, selected_date=None):
    """Get predictions for a specific user and optionally a specific date"""
    predictions_df = load_predictions()
    user_predictions = predictions_df[predictions_df['username'] == username]
    
    if selected_date:
        user_predictions = user_predictions[user_predictions['date'] == selected_date]
    
    return user_predictions.sort_values('date', ascending=False)

@app.before_request
def before_request():
    init_excel_files()

@app.route("/")
def index():
    if 'username' in session:
        return redirect(url_for('dashboard'))
    return render_template("index.html")

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]
        
        users_df = load_users()
        user = users_df[users_df['username'] == username]
        
        if not user.empty and check_password_hash(user.iloc[0]['password_hash'], password):
            session['username'] = username
            flash('Login successful!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid username or password!', 'error')
    
    return render_template("login.html")

@app.route("/signup", methods=["GET", "POST"])
def signup():
    if request.method == "POST":
        username = request.form["username"]
        email = request.form["email"]
        password = request.form["password"]
        confirm_password = request.form["confirm_password"]
        
        if password != confirm_password:
            flash('Passwords do not match!', 'error')
            return render_template("signup.html")
        
        users_df = load_users()
        if not users_df[users_df['username'] == username].empty:
            flash('Username already exists!', 'error')
            return render_template("signup.html")
        
        if not users_df[users_df['email'] == email].empty:
            flash('Email already exists!', 'error')
            return render_template("signup.html")
        
        save_user(username, email, password)
        flash('Account created successfully! Please login.', 'success')
        return redirect(url_for('login'))
    
    return render_template("signup.html")

@app.route("/logout")
def logout():
    session.pop('username', None)
    flash('You have been logged out.', 'info')
    return redirect(url_for('index'))

@app.route("/dashboard")
def dashboard():
    if 'username' not in session:
        return redirect(url_for('login'))
    
    username = session['username']
    recent_predictions = get_user_predictions(username).head(5)
    
    return render_template("dashboard.html", username=username, recent_predictions=recent_predictions)

@app.route("/history")
def history():
    if 'username' not in session:
        return redirect(url_for('login'))
    
    username = session['username']
    selected_date = request.args.get('date')
    
    predictions = get_user_predictions(username, selected_date)
    
    # Get unique dates for calendar
    all_predictions = get_user_predictions(username)
    prediction_dates = all_predictions['date'].unique().tolist() if not all_predictions.empty else []
    
    return render_template("history.html", 
                         username=username, 
                         predictions=predictions, 
                         selected_date=selected_date,
                         prediction_dates=prediction_dates)

@app.route("/breast-cancer", methods=["GET", "POST"])
def breast():
    if 'username' not in session:
        return redirect(url_for('login'))
    
    prob = None
    if request.method == "POST":
        features = [
            float(request.form["concavity_worst"]),
            float(request.form["perimeter_worst"]),
            float(request.form["concave_points_worst"]),
            float(request.form["radius_worst"]),
            float(request.form["radius_mean"]),
            float(request.form["area_worst"]),
            float(request.form["area_mean"]),
            float(request.form["perimeter_mean"])
        ]
        prob = breast_model.predict_proba([features])[0][1]
        
        # Save prediction
        input_data = {
            "concavity_worst": features[0],
            "perimeter_worst": features[1],
            "concave_points_worst": features[2],
            "radius_worst": features[3],
            "radius_mean": features[4],
            "area_worst": features[5],
            "area_mean": features[6],
            "perimeter_mean": features[7]
        }
        save_prediction(session['username'], 'Breast Cancer', input_data, prob)
        
    return render_template("breast-cancer.html", prob=prob)

@app.route("/cervical", methods=["GET", "POST"])
def cervical():
    if 'username' not in session:
        return redirect(url_for('login'))
    
    prob = None
    if request.method == "POST":
        features = [
            float(request.form["Age"]),
            float(request.form["Number_of_sexual_partners"]),
            float(request.form["First_sexual_intercourse"]),
            int(request.form["STDs_HPV"]),
            int(request.form["Dx_HPV"]),
            int(request.form["STDs"]),
            float(request.form["STDs_Number_of_diagnosis"]),
            float(request.form["STDs_Time_since_last_diagnosis"])
        ]
        prob = cervical_model.predict_proba([features])[0][1]
        
        # Save prediction
        input_data = {
            "Age": features[0],
            "Number_of_sexual_partners": features[1],
            "First_sexual_intercourse": features[2],
            "STDs_HPV": features[3],
            "Dx_HPV": features[4],
            "STDs": features[5],
            "STDs_Number_of_diagnosis": features[6],
            "STDs_Time_since_last_diagnosis": features[7]
        }
        save_prediction(session['username'], 'Cervical Cancer', input_data, prob)
        
    return render_template("cervical.html", prob=prob)

@app.route("/colorectal", methods=["GET", "POST"])
def colorectal():
    if 'username' not in session:
        return redirect(url_for('login'))
    
    prob = None
    if request.method == "POST":
        age = float(request.form["Age"])
        bmi = float(request.form["BMI"])
        data = crc_sample.copy()
        data["Age"] = age
        data["BMI"] = bmi
        prob = colorectal_model.predict_proba(data)[0][1]
        
        # Save prediction
        input_data = {
            "Age": age,
            "BMI": bmi
        }
        save_prediction(session['username'], 'Colorectal Cancer', input_data, prob)
        
    return render_template("colorectal.html", prob=prob)

if __name__ == "__main__":
    app.run(debug=True)